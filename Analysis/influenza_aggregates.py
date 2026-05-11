"""
influenza_aggregates.py
========================

Reproducible aggregation of the integrated state-year dataset for the
U.S. Influenza Staffing Analysis. Computes the four headline statistics
from the case study writeup directly from the Excel source.

Usage:
    python3 Analysis/influenza_aggregates.py

Requires:
    openpyxl, numpy
"""

import collections
import statistics
from pathlib import Path

import numpy as np
import openpyxl

# --------------------------------------------------------------------------
# Paths
# --------------------------------------------------------------------------

REPO_ROOT = Path(__file__).resolve().parent.parent
INTEGRATED_XLSX = REPO_ROOT / "Data" / "1_7_Data_Integration_Akeel_Alramadhan.xlsx"
RAW_DEATHS_XLSX = REPO_ROOT / "Data" / "CDC_Influenza_Deaths_edited.xlsx"


# --------------------------------------------------------------------------
# Step 1 — Load the integrated state-year master (the analytical table)
# --------------------------------------------------------------------------

def load_integrated():
    """Return list of dicts with state-year-level deaths, population, rate."""
    wb = openpyxl.load_workbook(INTEGRATED_XLSX, data_only=True, read_only=True)
    ws = wb["Integrated"]
    rows = []
    for r in ws.iter_rows(values_only=True, min_row=2):
        state, year, population, key, deaths, rate = r
        if state is None or year is None or deaths is None:
            continue
        rows.append({
            "state": state,
            "year": int(year),
            "population": int(population),
            "deaths": int(deaths),
            "rate": float(rate) if rate is not None else None,
        })
    return rows


# --------------------------------------------------------------------------
# Step 2 — Year-over-year national mortality rate
# --------------------------------------------------------------------------

def yearly_summary(rows):
    deaths = collections.defaultdict(int)
    pop = collections.defaultdict(int)
    for r in rows:
        deaths[r["year"]] += r["deaths"]
        pop[r["year"]] += r["population"]
    summary = {}
    for y in sorted(deaths):
        summary[y] = {
            "deaths": deaths[y],
            "population": pop[y],
            "rate_per_100k": deaths[y] / pop[y] * 100_000,
        }
    return summary


# --------------------------------------------------------------------------
# Step 3 — Population vs deaths correlation (state-year grain)
# --------------------------------------------------------------------------

def pop_deaths_correlation(rows):
    pops = np.array([r["population"] for r in rows])
    deaths = np.array([r["deaths"] for r in rows])
    return float(np.corrcoef(pops, deaths)[0, 1])


# --------------------------------------------------------------------------
# Step 4 — Top-10 states by total deaths vs by avg mortality rate
# --------------------------------------------------------------------------

def top_10_rankings(rows):
    deaths_by_state = collections.defaultdict(int)
    pop_by_state = collections.defaultdict(int)
    for r in rows:
        deaths_by_state[r["state"]] += r["deaths"]
        pop_by_state[r["state"]] += r["population"]

    by_count = sorted(deaths_by_state.items(), key=lambda x: -x[1])[:10]
    by_rate = sorted(
        deaths_by_state.keys(),
        key=lambda s: -(deaths_by_state[s] / pop_by_state[s] * 100_000),
    )[:10]
    by_rate = [
        (s, deaths_by_state[s] / pop_by_state[s] * 100_000) for s in by_rate
    ]
    return by_count, by_rate


# --------------------------------------------------------------------------
# Step 5 — Age-group breakdown (from the raw CDC table)
# --------------------------------------------------------------------------

US_STATE_NAMES = {
    "Alabama","Alaska","Arizona","Arkansas","California","Colorado","Connecticut",
    "Delaware","District of Columbia","Florida","Georgia","Hawaii","Idaho","Illinois",
    "Indiana","Iowa","Kansas","Kentucky","Louisiana","Maine","Maryland","Massachusetts",
    "Michigan","Minnesota","Mississippi","Missouri","Montana","Nebraska","Nevada",
    "New Hampshire","New Jersey","New Mexico","New York","North Carolina","North Dakota",
    "Ohio","Oklahoma","Oregon","Pennsylvania","Rhode Island","South Carolina","South Dakota",
    "Tennessee","Texas","Utah","Vermont","Virginia","Washington","West Virginia",
    "Wisconsin","Wyoming",
}
US_STATE_ABBR = {
    "AK","AL","AR","AZ","CA","CO","CT","DE","FL","GA","HI","IA","ID","IL","IN","KS",
    "KY","LA","MA","MD","ME","MI","MN","MO","MS","MT","NC","ND","NE","NH","NJ","NM",
    "NV","NY","OH","OK","OR","PA","RI","SC","SD","TN","TX","UT","VA","VT","WA","WI",
    "WV","WY",
}
US_ALL = US_STATE_NAMES | US_STATE_ABBR


def age_breakdown():
    wb = openpyxl.load_workbook(RAW_DEATHS_XLSX, data_only=True, read_only=True)
    ws = wb["Month and Age"]
    age_deaths = collections.defaultdict(int)
    for row in ws.iter_rows(values_only=True, min_row=2):
        state, _, year, _, _, age_group, _, deaths = row
        if state not in US_ALL:
            continue
        if year == 20133:  # data-entry typo in source file
            continue
        if deaths in (None, "Suppressed"):
            continue
        try:
            age_deaths[age_group] += int(deaths)
        except (TypeError, ValueError):
            continue
    return dict(age_deaths)


# --------------------------------------------------------------------------
# Main
# --------------------------------------------------------------------------

def main():
    print("=" * 70)
    print("U.S. Influenza Staffing Analysis — Aggregate Statistics")
    print("=" * 70)

    rows = load_integrated()
    print(f"\nLoaded {len(rows)} state-year observations across "
          f"{len(set(r['state'] for r in rows))} jurisdictions and "
          f"{len(set(r['year'] for r in rows))} years.")

    # --- Yearly ---
    print("\n--- National mortality, by year ---")
    print(f"{'Year':<6}{'Deaths':>12}{'Population':>16}{'Rate /100k':>14}")
    summary = yearly_summary(rows)
    for y, s in summary.items():
        print(f"{y:<6}{s['deaths']:>12,}{s['population']:>16,}"
              f"{s['rate_per_100k']:>14.2f}")
    rates = [s["rate_per_100k"] for s in summary.values()]
    print(f"\nNational rate range: {min(rates):.2f}–{max(rates):.2f} per 100k")
    print(f"National rate mean:  {statistics.mean(rates):.2f} per 100k "
          f"(≈ {statistics.mean(rates)/1000:.4f}%)")

    # --- Correlation ---
    r = pop_deaths_correlation(rows)
    print(f"\n--- Population vs deaths correlation (state-year grain) ---")
    print(f"Pearson r = {r:.3f}")

    # --- Rankings ---
    by_count, by_rate = top_10_rankings(rows)
    print(f"\n--- Top 10 states by total deaths ---")
    for i, (s, d) in enumerate(by_count, 1):
        print(f"  {i:>2}. {s:<20}{d:>10,}")
    print(f"\n--- Top 10 states by mortality rate (per 100k) ---")
    for i, (s, rate) in enumerate(by_rate, 1):
        print(f"  {i:>2}. {s:<20}{rate:>10.2f}")

    overlap = set(s for s, _ in by_count) & set(s for s, _ in by_rate)
    print(f"\nStates appearing on both top-10 lists: {sorted(overlap)}")

    # --- Age ---
    age = age_breakdown()
    print(f"\n--- Deaths by 10-year age group ---")
    total_age = sum(age.values())
    order = sorted(age.items(), key=lambda x: -x[1])
    for ag, d in order:
        print(f"  {ag:<14}{d:>10,}  ({d/total_age*100:>5.1f}%)")

    age_85plus = age.get("85+ years", 0)
    age_75plus = age_85plus + age.get("75-84 years", 0)
    age_65plus = age_75plus + age.get("65-74 years", 0)
    print(f"\n85+ alone:   {age_85plus:>10,}  ({age_85plus/total_age*100:.1f}%)")
    print(f"75+ combined:{age_75plus:>10,}  ({age_75plus/total_age*100:.1f}%)")
    print(f"65+ combined:{age_65plus:>10,}  ({age_65plus/total_age*100:.1f}%)")

    print("\n" + "=" * 70)
    print("Done.")
    print("=" * 70)


if __name__ == "__main__":
    main()
